from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
import datetime, json, io

from .models import Candidate, Application, Payment, Invoice, Document, CourierEntry, LegalRecord, B2BClient
from .forms import CandidateForm, ApplicationForm, PaymentForm, InvoiceForm, DocumentForm, CourierForm, LegalForm, B2BForm


# ─── AUTH ─────────────────────────────────────────────
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username','').strip()
        password = request.POST.get('password','')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    return render(request, 'crm/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ─── DASHBOARD ────────────────────────────────────────
@login_required
def dashboard(request):
    total_candidates = Candidate.objects.count()
    boarded = Application.objects.filter(status='Boarded').count()
    active = Application.objects.exclude(status__in=['Boarded','Cancelled','Refunded']).count()
    revenue = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
    pending_balance = sum(a.balance for a in Application.objects.all())
    open_invoices = Invoice.objects.filter(status='Pending').count()
    recent_apps = Application.objects.select_related('candidate').order_by('-created_at')[:8]
    followups_today = Candidate.objects.filter(followup_date=datetime.date.today()).count()

    context = {
        'total_candidates': total_candidates,
        'boarded': boarded,
        'active': active,
        'revenue': revenue,
        'pending_balance': pending_balance,
        'open_invoices': open_invoices,
        'recent_apps': recent_apps,
        'followups_today': followups_today,
        'page': 'dashboard',
    }
    return render(request, 'crm/dashboard.html', context)


# ─── CANDIDATES ───────────────────────────────────────
@login_required
def candidates(request):
    q = request.GET.get('q','')
    qs = Candidate.objects.all()
    if q:
        qs = qs.filter(Q(name__icontains=q)|Q(phone__icontains=q)|Q(passport_number__icontains=q)|Q(country__icontains=q)|Q(position__icontains=q))
    form = CandidateForm()
    return render(request, 'crm/candidates.html', {'candidates': qs, 'form': form, 'q': q, 'page': 'candidates'})


@login_required
def candidate_add(request):
    if request.method == 'POST':
        form = CandidateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Candidate added successfully!')
        else:
            messages.error(request, 'Error: ' + str(form.errors))
    return redirect('candidates')


@login_required
def candidate_edit(request, pk):
    c = get_object_or_404(Candidate, pk=pk)
    if request.method == 'POST':
        form = CandidateForm(request.POST, request.FILES, instance=c)
        if form.is_valid():
            form.save()
            messages.success(request, 'Candidate updated!')
            return redirect('candidates')
    else:
        form = CandidateForm(instance=c)
    return render(request, 'crm/candidate_edit.html', {'form': form, 'candidate': c, 'page': 'candidates'})


@login_required
def candidate_view(request, pk):
    c = get_object_or_404(Candidate, pk=pk)
    payments = c.payments.all()
    applications = c.applications.all()
    documents = c.documents.all()
    legal = c.legal_records.first()
    total_paid = payments.aggregate(t=Sum('amount'))['t'] or 0
    return render(request, 'crm/candidate_view.html', {
        'candidate': c, 'payments': payments, 'applications': applications,
        'documents': documents, 'legal': legal, 'total_paid': total_paid, 'page': 'candidates'
    })


@login_required
def candidate_delete(request, pk):
    get_object_or_404(Candidate, pk=pk).delete()
    messages.success(request, 'Candidate deleted.')
    return redirect('candidates')


# ─── APPLICATIONS ─────────────────────────────────────
@login_required
def applications(request):
    status_filter = request.GET.get('status','')
    q = request.GET.get('q','')
    qs = Application.objects.select_related('candidate').all()
    if status_filter:
        qs = qs.filter(status=status_filter)
    if q:
        qs = qs.filter(Q(candidate__name__icontains=q)|Q(country__icontains=q)|Q(job_category__icontains=q))
    form = ApplicationForm()
    statuses = [s[0] for s in Application.STATUS_CHOICES]
    return render(request, 'crm/applications.html', {
        'applications': qs, 'form': form, 'statuses': statuses,
        'status_filter': status_filter, 'q': q, 'page': 'applications'
    })


@login_required
def application_add(request):
    if request.method == 'POST':
        form = ApplicationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Application saved!')
        else:
            messages.error(request, 'Error: ' + str(form.errors))
    return redirect('applications')


@login_required
def application_delete(request, pk):
    get_object_or_404(Application, pk=pk).delete()
    messages.success(request, 'Application deleted.')
    return redirect('applications')


# ─── DOCUMENTS ────────────────────────────────────────
@login_required
def documents(request):
    q = request.GET.get('q','')
    qs = Document.objects.select_related('candidate').all()
    if q:
        qs = qs.filter(Q(candidate__name__icontains=q)|Q(doc_type__icontains=q))
    form = DocumentForm()
    return render(request, 'crm/documents.html', {'documents': qs, 'form': form, 'q': q, 'page': 'documents'})


@login_required
def document_add(request):
    if request.method == 'POST':
        form = DocumentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Document record added!')
        else:
            messages.error(request, str(form.errors))
    return redirect('documents')


@login_required
def document_delete(request, pk):
    get_object_or_404(Document, pk=pk).delete()
    messages.success(request, 'Deleted.')
    return redirect('documents')


# ─── COURIER ──────────────────────────────────────────
@login_required
def courier(request):
    q = request.GET.get('q','')
    qs = CourierEntry.objects.select_related('candidate').all()
    if q:
        qs = qs.filter(Q(candidate__name__icontains=q)|Q(tracking_number__icontains=q)|Q(courier_company__icontains=q))
    form = CourierForm()
    return render(request, 'crm/courier.html', {'entries': qs, 'form': form, 'q': q, 'page': 'courier'})


@login_required
def courier_add(request):
    if request.method == 'POST':
        form = CourierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Courier entry added!')
        else:
            messages.error(request, str(form.errors))
    return redirect('courier')


@login_required
def courier_delete(request, pk):
    get_object_or_404(CourierEntry, pk=pk).delete()
    messages.success(request, 'Deleted.')
    return redirect('courier')


# ─── PAYMENTS ─────────────────────────────────────────
@login_required
def payments(request):
    q = request.GET.get('q','')
    qs = Payment.objects.select_related('candidate').all()
    if q:
        qs = qs.filter(Q(candidate__name__icontains=q)|Q(purpose__icontains=q)|Q(mode__icontains=q))
    form = PaymentForm()
    total = qs.aggregate(t=Sum('amount'))['t'] or 0
    return render(request, 'crm/payments.html', {'payments': qs, 'form': form, 'q': q, 'total': total, 'page': 'payments'})


@login_required
def payment_add(request):
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Payment recorded!')
        else:
            messages.error(request, str(form.errors))
    return redirect('payments')


@login_required
def payment_delete(request, pk):
    get_object_or_404(Payment, pk=pk).delete()
    messages.success(request, 'Deleted.')
    return redirect('payments')


# ─── INVOICES ─────────────────────────────────────────
@login_required
def invoices(request):
    q = request.GET.get('q','')
    qs = Invoice.objects.select_related('candidate').all()
    if q:
        qs = qs.filter(Q(invoice_number__icontains=q)|Q(candidate__name__icontains=q)|Q(domain__icontains=q))
    form = InvoiceForm()
    return render(request, 'crm/invoices.html', {'invoices': qs, 'form': form, 'q': q, 'page': 'invoices'})


@login_required
def invoice_add(request):
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Invoice created!')
        else:
            messages.error(request, str(form.errors))
    return redirect('invoices')


@login_required
def invoice_mark_paid(request, pk):
    inv = get_object_or_404(Invoice, pk=pk)
    inv.status = 'Cleared'
    inv.save()
    messages.success(request, 'Invoice marked as cleared!')
    return redirect('invoices')


@login_required
def invoice_delete(request, pk):
    get_object_or_404(Invoice, pk=pk).delete()
    messages.success(request, 'Deleted.')
    return redirect('invoices')


# ─── LEGAL ────────────────────────────────────────────
@login_required
def legal(request):
    q = request.GET.get('q','')
    qs = LegalRecord.objects.select_related('candidate').all()
    if q:
        qs = qs.filter(candidate__name__icontains=q)
    form = LegalForm()
    return render(request, 'crm/legal.html', {'records': qs, 'form': form, 'q': q, 'page': 'legal'})


@login_required
def legal_add(request):
    if request.method == 'POST':
        form = LegalForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Legal record saved!')
        else:
            messages.error(request, str(form.errors))
    return redirect('legal')


@login_required
def legal_delete(request, pk):
    get_object_or_404(LegalRecord, pk=pk).delete()
    messages.success(request, 'Deleted.')
    return redirect('legal')


# ─── B2B ──────────────────────────────────────────────
@login_required
def b2b(request):
    q = request.GET.get('q','')
    qs = B2BClient.objects.all()
    if q:
        qs = qs.filter(Q(company_name__icontains=q)|Q(contact_person__icontains=q)|Q(country__icontains=q))
    form = B2BForm()
    return render(request, 'crm/b2b.html', {'clients': qs, 'form': form, 'q': q, 'page': 'b2b'})


@login_required
def b2b_add(request):
    if request.method == 'POST':
        form = B2BForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'B2B client added!')
        else:
            messages.error(request, str(form.errors))
    return redirect('b2b')


@login_required
def b2b_delete(request, pk):
    get_object_or_404(B2BClient, pk=pk).delete()
    messages.success(request, 'Deleted.')
    return redirect('b2b')


# ─── EXCEL EXPORT ─────────────────────────────────────
@login_required
def export_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a237e")

    def make_sheet(wb, title, headers, rows, first=False):
        ws = wb.active if first else wb.create_sheet(title)
        ws.title = title
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        return ws

    # Candidates sheet
    cands = Candidate.objects.all()
    make_sheet(wb, 'Candidates',
        ['#','Name','Phone','Passport No','Country','Position','Experience','Status',
         'Medical','Visa','Ticket','Payment(₹)','Resume','Passport Copy','Offer Letter',
         'Travel Date','Reference','Remarks','Follow-up Date','Staff','Reg Date','Type'],
        [[i+1, c.name, c.phone, c.passport_number, c.country, c.position, c.experience,
          c.status, c.medical_status, c.visa_status, c.ticket_status, float(c.payment),
          c.resume_status, c.passport_copy_status, c.offer_letter_status,
          str(c.travel_date or ''), c.reference, c.remarks,
          str(c.followup_date or ''), c.staff_name, str(c.registration_date or ''), c.candidate_type]
         for i, c in enumerate(cands)], first=True)

    # Applications
    apps = Application.objects.select_related('candidate').all()
    make_sheet(wb, 'Applications',
        ['#','Candidate','Country','Job','Status','App Date','Board Date','Package(₹)','Paid(₹)','Balance(₹)','Remarks'],
        [[i+1, a.candidate.name, a.country, a.job_category, a.status,
          str(a.application_date or ''), str(a.boarding_date or ''),
          float(a.package_amount), float(a.amount_paid), float(a.balance), a.remarks]
         for i, a in enumerate(apps)])

    # Payments
    pays = Payment.objects.select_related('candidate').all()
    make_sheet(wb, 'Payments',
        ['#','Candidate','Amount(₹)','Mode','Purpose','Date','Reference','Notes'],
        [[i+1, p.candidate.name, float(p.amount), p.mode, p.purpose,
          str(p.date), p.reference, p.notes]
         for i, p in enumerate(pays)])

    # Invoices
    invs = Invoice.objects.select_related('candidate').all()
    make_sheet(wb, 'Invoices',
        ['#','Invoice No','Domain','Candidate','Amount(₹)','Date','Status','Description'],
        [[i+1, inv.invoice_number, inv.domain,
          inv.candidate.name if inv.candidate else '',
          float(inv.amount), str(inv.date), inv.status, inv.description]
         for i, inv in enumerate(invs)])

    # Documents
    docs = Document.objects.select_related('candidate').all()
    make_sheet(wb, 'Documents',
        ['#','Candidate','Doc Type','Original','Attested','Verified','With Company','Notes'],
        [[i+1, d.candidate.name, d.doc_type,
          'Yes' if d.is_original else 'No',
          'Yes' if d.is_attested else 'No',
          'Yes' if d.is_verified else 'No',
          'Yes' if d.with_company else 'No', d.notes]
         for i, d in enumerate(docs)])

    # Courier
    couriers = CourierEntry.objects.select_related('candidate').all()
    make_sheet(wb, 'Courier',
        ['#','Type','Candidate','From','To','Company','Tracking No','Date','Cost(₹)','Contents','Handled By'],
        [[i+1, cr.entry_type,
          cr.candidate.name if cr.candidate else '',
          cr.from_person, cr.to_person, cr.courier_company, cr.tracking_number,
          str(cr.date), float(cr.cost), cr.contents, cr.handled_by]
         for i, cr in enumerate(couriers)])

    # Legal
    legals = LegalRecord.objects.select_related('candidate').all()
    make_sheet(wb, 'Legal',
        ['#','Candidate','Agreement','SOP Shared','Terms Accepted','Refund Eligible','Refund Amt(₹)','Refund Date','Cancellation','Notes'],
        [[i+1, l.candidate.name,
          'Yes' if l.agreement_signed else 'No',
          'Yes' if l.sop_shared else 'No',
          'Yes' if l.terms_accepted else 'No',
          'Yes' if l.refund_eligible else 'No',
          float(l.refund_amount), str(l.refund_date or ''), l.cancellation_status, l.notes]
         for i, l in enumerate(legals)])

    # B2B
    b2bs = B2BClient.objects.all()
    make_sheet(wb, 'B2B Clients',
        ['#','Company','Contact','Phone','Email','Country','Contract Status','Notes'],
        [[i+1, b.company_name, b.contact_person, b.phone, b.email,
          b.country, b.contract_status, b.notes]
         for i, b in enumerate(b2bs)])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    today = datetime.date.today().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="ImmigrateCRM_Export_{today}.xlsx"'
    wb.save(response)
    return response


# ─── EXCEL IMPORT ─────────────────────────────────────
@login_required
def import_excel(request):
    if request.method != 'POST' or 'excel_file' not in request.FILES:
        messages.error(request, 'No file uploaded.')
        return redirect('candidates')

    try:
        import openpyxl
    except ImportError:
        messages.error(request, 'openpyxl not installed.')
        return redirect('candidates')

    f = request.FILES['excel_file']
    wb = openpyxl.load_workbook(f)
    imported = 0

    if 'Candidates' in wb.sheetnames:
        ws = wb['Candidates']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[1]:
                continue
            try:
                Candidate.objects.get_or_create(
                    name=str(row[1] or '').strip(),
                    phone=str(row[2] or '').strip(),
                    defaults={
                        'passport_number': str(row[3] or ''),
                        'country': str(row[4] or ''),
                        'position': str(row[5] or ''),
                        'experience': str(row[6] or ''),
                        'status': str(row[7] or 'New'),
                        'medical_status': str(row[8] or 'Pending'),
                        'visa_status': str(row[9] or 'Not Applied'),
                        'ticket_status': str(row[10] or 'Not Booked'),
                        'payment': float(row[11] or 0),
                        'resume_status': str(row[12] or 'Not Received'),
                        'passport_copy_status': str(row[13] or 'Not Received'),
                        'offer_letter_status': str(row[14] or 'Not Received'),
                        'reference': str(row[16] or ''),
                        'remarks': str(row[17] or ''),
                        'staff_name': str(row[19] or ''),
                        'candidate_type': str(row[21] or 'New'),
                    }
                )
                imported += 1
            except Exception:
                pass

    messages.success(request, f'Imported {imported} candidates from Excel!')
    return redirect('candidates')
